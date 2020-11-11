import openpyxl
import panda

# wb = openpyxl.load_workbook('example.xlsx')#参数中填路径（r/D:xxx.xlsx）
# sheet_name = wb.sheetnames#获取表格工作表名称
# print(wb.worksheets)
# sheet_a = wb[sheet_name[0]] #工作表A
# lis = []
# for coloum in (list(sheet_a.columns)[-1]):
#     date = col.value for coloum in coloum:
# for column in sheet_a.columns:
#
#     for cell in column:
#         lis.append(cell.value)
# print(lis)
# #
# # print('*'*50)
# # print(sheet_a.max_row)
# # print(sheet_a.max_column)
# # print(sheet_a.cell(row=2,column=3).value)#根据第几行第几列定位单元格
# # for i in range(1,8,2):#设定行数范围，按2的步长拿出单元格数据
# #     print(i,sheet_a.cell(row=i,column=3).value)
# #
# # a1 = sheet_a['B1'].value #值
# # a1_1 = sheet_a['B1'].row#行
# # a1_2 = sheet_a['B1'].column#列
# # a1_xy = sheet_a['B1'].coordinate#单元格坐标名称


from openpyxl import load_workbook
people = load_workbook('people.xlsx')
print(people.sheetnames)
people.create_sheet(title='jj',index=0)#创建表
print(people.sheetnames)
table_jj = people['Sheet1']

table_jj.title = 'ss'
print(people.sheetnames)
print(table_jj['A6'].value) #定位单元格
#将第1列改为百分比
table_jj.cell(row=4,column=1).number_format='0.00%'
table_jj.cell(row=5,column=1).number_format='0.00%'
table_jj.cell(row=6,column=1).number_format='0.00%'
table_jj.cell(row=7,column=1).number_format='0.00%'

#将第2列改为小数点后两位
table_jj.cell(row=4,column=2).number_format='0.00'
table_jj.cell(row=5,column=2).number_format='0.00'
table_jj.cell(row=6,column=2).number_format='0.00'
table_jj.cell(row=7,column=2).number_format='0.00'

#将第3列改为带货币符号
table_jj.cell(row=4,column=2).number_format='"$"#,###'
table_jj.cell(row=5,column=2).number_format='"$"#,###'
table_jj.cell(row=6,column=2).number_format='"$"#,###'
table_jj.cell(row=7,column=2).number_format='"$"#,###'

#将第4列改为年月日时分秒格式
table_jj.cell(row=4,column=4).number_format = 'yyyy-MM-dd HH:mm:ss'
table_jj.cell(row=5,column=4).number_format = 'yyyy-MM-dd HH:mm:ss'
table_jj.cell(row=6,column=4).number_format = 'yyyy-MM-dd HH:mm:ss'
table_jj.cell(row=7,column=4).number_format = 'yyyy-MM-dd HH:mm:ss'

#copy表
new_table = people.copy_worksheet(table_jj)

#将复制后的工作表名称用红色填充
new_table.title = 'The_red'
new_table.sheet_properties.tabColor = 'FF0000'

people.save('people.xlsx')
